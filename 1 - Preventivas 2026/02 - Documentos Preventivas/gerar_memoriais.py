# gerar_memoriais.py
# Lê as planilhas de previsão orçamentária de cada contrato e gera:
#   - NNNN_CIDADE_itens.json  (dado canônico para o AutoRelatório V5)
#   - NNNN_CIDADE_memorial_itens.html  (visualização interativa)
# Somente leitura — nenhum arquivo original é modificado.

import os, re, json, unicodedata
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))

# Mapeamento: parte do nome da pasta → (codigo_contrato, slug_cidade)
CONTRATOS = {
    "2056": ("2056", "DIVINOPOLIS"),
    "2057": ("2057", "VARGINHA"),
    "6122": ("6122", "MS"),
    "0908": ("0908", "SAO_PAULO"),
    "1507": ("1507", "CUIABA"),
    "2626": ("2626", "SALINAS"),
    "2627": ("2627", "VALADARES"),
    "3575": ("3575", "TANGARA"),
    "1565": ("1565", "SJRP"),
}

# Normalização de unidades de medida
UN_NORM = {
    "m²": "m²", "m2": "m²", "M²": "m²", "M2": "m²",
    "m³": "m³", "m3": "m³", "M³": "m³", "M3": "m³",
    "m": "m", "M": "m",
    "un": "un", "UN": "un", "un ": "un",
    "cj": "cj", "CJ": "cj",
    "km": "km", "Km": "km", "KM": "km",
    "m²/mês": "m²/mês", "m2/mês": "m²/mês", "m²/mes": "m²/mês",
    "m/mês": "m/mês", "m/mes": "m/mês",
    "mês": "mês", "mes": "mês",
    "H": "H", "h": "H",
    "m²/m²s": "m²/mês",
}

def norm_un(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    # Substitui 'ê' e variantes
    s = s.replace("\xea", "ê")
    return UN_NORM.get(s, s)

def limpar(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s in ("#N/A", "0", "0.0"):
        return s
    return s

def e_item(codigo_str):
    """Retorna True se o código tem ponto (ex: '2.12', '29.11')"""
    s = str(codigo_str).strip()
    return bool(re.match(r"^\d+\.\d+$", s))

def e_secao(codigo_str, descricao_str):
    """Retorna True se parece ser uma linha de seção (ex: '2', 'IMPLANTAÇÃO')"""
    s = str(codigo_str).strip()
    return bool(re.match(r"^\d+$", s)) and descricao_str

def detectar_aba_valores(wb):
    for nome in wb.sheetnames:
        n = nome.strip().lower()
        if "valores unit" in n:
            return wb[nome]
        if n == "valores":
            return wb[nome]
    return None

def detectar_aba_prefixos(wb):
    for nome in wb.sheetnames:
        if "prefixo" in nome.lower():
            return wb[nome]
    return None

def extrair_cabecalho(ws):
    """Extrai contrato, lote, desconto, proponente da aba de valores."""
    contrato_num = ""
    lote = ""
    desconto = None
    proponente = ""
    bdi = None
    for row in ws.iter_rows(values_only=True, max_row=10):
        row_vals = [str(c).strip() if c is not None else "" for c in row]
        linha = " ".join(row_vals)
        # Contrato
        m = re.search(r"(\d{4}\.7421\.\d{4})", linha)
        if m:
            contrato_num = m.group(1)
        # Lote
        for i, v in enumerate(row_vals):
            if v.upper() == "LOTE:" and i + 1 < len(row_vals) and row_vals[i+1]:
                lote = row_vals[i+1]
            if v.upper() == "DESCONTO:" and i + 1 < len(row_vals):
                try:
                    desconto = float(row_vals[i+1])
                except:
                    pass
        # Proponente (linha com CNPJ)
        if "CNPJ" in linha and not proponente:
            for v in row_vals:
                if v and "CNPJ" not in v and "MAFFENG" in v.upper():
                    proponente = v
                    break
    return contrato_num, lote, desconto, proponente

def extrair_bdi(wb):
    for nome in wb.sheetnames:
        if "bdi" in nome.lower():
            ws = wb[nome]
            for row in ws.iter_rows(values_only=True):
                row_vals = [str(c).strip() if c is not None else "" for c in row]
                for i, v in enumerate(row_vals):
                    if "BDI SERVI" in v.upper() and i + 1 < len(row_vals):
                        try:
                            return float(row_vals[i+1])
                        except:
                            pass
    return None

def extrair_itens(ws):
    """Extrai seções e itens da aba de valores."""
    secoes = []
    secao_atual = None
    # Detecta colunas: ITEM | DESCRIÇÃO | QTDE | UN
    col_item = 0
    col_desc = 1
    col_qtde = 2
    col_un = 3

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not any(c is not None for c in cells):
            continue

        # Busca valor da coluna ITEM
        raw_item = cells[col_item] if len(cells) > col_item else None
        raw_desc = cells[col_desc] if len(cells) > col_desc else None
        raw_qtde = cells[col_qtde] if len(cells) > col_qtde else None
        raw_un   = cells[col_un]   if len(cells) > col_un   else None

        if raw_item is None:
            continue

        item_str = str(raw_item).strip()
        desc_str = str(raw_desc).strip() if raw_desc else ""

        # Ignora cabeçalhos e lixo
        if item_str.upper() in ("ITEM", "", "A INTERESSADA", "CONSERVAÇÃO", "PROPONENTE"):
            continue
        if "#N/A" in item_str or "#N/A" in desc_str:
            continue
        if item_str.upper() in ("DISEC", "CESUP", "SUPRIMENTO"):
            continue

        # Linha de seção
        if e_secao(item_str, desc_str):
            secao_atual = {"codigo": item_str, "nome": desc_str, "itens": []}
            secoes.append(secao_atual)
            continue

        # Linha de item real
        if e_item(item_str) and desc_str:
            un_raw = str(raw_un).strip() if raw_un is not None else ""
            un = norm_un(un_raw) if un_raw and un_raw not in ("#N/A", "0") else None
            if un is None:
                continue
            item_obj = {
                "item": item_str,
                "descricao": desc_str,
                "unidade": un
            }
            if secao_atual is None:
                secao_atual = {"codigo": "?", "nome": "GERAL", "itens": []}
                secoes.append(secao_atual)
            secao_atual["itens"].append(item_obj)

    return secoes

def extrair_prefixos(ws):
    prefixos = []
    cabecalho_passado = False
    idx_pref = 0
    idx_nome = 1
    idx_km   = None
    idx_end  = None
    idx_cid  = None

    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(c for c in cells):
            continue

        if not cabecalho_passado:
            linha = " ".join(cells).upper()
            if "PREFIXO" not in linha and "PREF" not in linha:
                continue
            cabecalho_passado = True
            upper = [c.upper() for c in cells]

            # Detecta coluna SB (pula ela para nome)
            has_sb = any(c == "SB" for c in upper)

            # Detecta coluna de KM
            for i, c in enumerate(upper):
                if "DIST" in c or "KM" in c:
                    idx_km = i
                    break

            # Detecta coluna de endereço
            for i, c in enumerate(upper):
                if "ENDERE" in c:
                    idx_end = i
                    break

            # Detecta coluna de cidade/município (não confundir com coluna de km)
            for i, c in enumerate(upper):
                if ("MUNIC" in c or "CIDADE" in c) and i != idx_km:
                    idx_cid = i
                    break

            # Nome da agência/dependência
            if has_sb:
                idx_nome = 2  # PREFIXO(0) SB(1) NOME(2)
            else:
                # Procura coluna de nome/dependência
                for i, c in enumerate(upper):
                    if any(k in c for k in ("DEPEND", "NOME", "AGENCI", "AGÊN")):
                        idx_nome = i
                        break
                else:
                    idx_nome = 1
            continue

        col_pref = cells[idx_pref] if len(cells) > idx_pref else ""
        if not col_pref or col_pref.upper() in ("PREFIXO", "PREFIXO/SB", ""):
            continue
        # Aceita "0046/00" (Cuiabá) e números puros
        pref_num = col_pref.split("/")[0].strip()
        try:
            int(float(pref_num))
        except:
            continue
        col_pref = pref_num.zfill(4)  # padroniza para 4 dígitos (ex: 76 → 0076)

        col_nome = cells[idx_nome] if idx_nome is not None and len(cells) > idx_nome else ""
        col_km   = cells[idx_km]   if idx_km  is not None and len(cells) > idx_km  else ""
        col_end  = cells[idx_end]  if idx_end  is not None and len(cells) > idx_end  else ""
        col_cid  = cells[idx_cid]  if idx_cid  is not None and len(cells) > idx_cid  else ""

        # Normaliza km: pode ser número float como "252.0"
        km_val = ""
        if col_km:
            try:
                km_val = str(int(float(col_km))) + " km"
            except:
                km_val = col_km

        # Limpa cidade: remove \n, descarta se numérica
        cidade_ok = col_cid.replace("\n", " ").replace("\r", "").strip()
        try:
            float(cidade_ok)
            cidade_ok = ""  # era número, não cidade
        except:
            pass
        # Suprime cidade só quando igual ao nome E há endereço (evita redundância)
        if cidade_ok.upper() == col_nome.upper() and col_end:
            cidade_ok = ""

        prefixos.append({
            "prefixo":    col_pref,
            "dependencia": col_nome,
            "km":         km_val,
            "endereco":   col_end,
            "cidade":     cidade_ok,
        })
    return prefixos

def montar_itens_por_unidade(secoes):
    mapa = {}
    for sec in secoes:
        for it in sec["itens"]:
            un = it["unidade"]
            if un not in mapa:
                mapa[un] = []
            mapa[un].append(it["item"])
    return mapa

def processar_contrato(xlsx_path, codigo, slug):
    print(f"\n  Processando {codigo} — {slug}")
    print(f"    Arquivo: {os.path.basename(xlsx_path)}")

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"    ERRO ao abrir: {e}")
        return None

    ws_val = detectar_aba_valores(wb)
    if not ws_val:
        print(f"    AVISO: Aba de valores não encontrada")
        wb.close()
        return None

    contrato_num, lote, desconto, proponente = extrair_cabecalho(ws_val)
    bdi = extrair_bdi(wb)
    secoes = extrair_itens(ws_val)
    itens_por_un = montar_itens_por_unidade(secoes)

    ws_pref = detectar_aba_prefixos(wb)
    prefixos = extrair_prefixos(ws_pref) if ws_pref else []

    wb.close()

    # Monta nome da cidade legível
    nome_cidade = slug.replace("_", " ")

    dados = {
        "codigo": codigo,
        "nome": nome_cidade,
        "contrato": contrato_num,
        "lote": lote,
        "desconto": round(desconto, 4) if desconto else None,
        "bdi": round(bdi, 4) if bdi else None,
        "proponente": proponente or "MAFFENG ENGENHARIA E MANUTENÇÃO",
        "secoes": secoes,
        "itens_por_unidade": itens_por_un,
        "prefixos": prefixos,
    }

    total_itens = sum(len(s["itens"]) for s in secoes)
    print(f"    Seções: {len(secoes)} | Itens: {total_itens} | Unidades: {list(itens_por_un.keys())} | Prefixos: {len(prefixos)}")
    return dados


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{TITULO}}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Roboto+Slab:wght@400;500;600&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
<style>
  /* ── TM Design System v3 — tokens canônicos ── */
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  :root {
    --TM-primary:             #C8541C;
    --TM-primary-hover:       #A6451A;
    --TM-primary-light:       #FBEDE3;
    --TM-primary-light-alpha: rgba(200,84,28,0.12);
    --TM-accent-graphite:     #2B2B2B;
    --tm-bg:                  #F5F4F1;
    --tm-bg-card:             #FFFFFF;
    --tm-bg-hover:            #ECEAE5;
    --tm-bg-input:            #FFFFFF;
    --tm-bg-sidebar:          #FAF9F6;
    --tm-bg-secondary:        #E8E5DE;
    --tm-bg-accent:           #FBEDE3;
    --tm-bg-muted:            #EFEDE8;
    --tm-text:                #1A1A1A;
    --tm-text-muted:          #5C5A55;
    --tm-text-subtle:         #8C8A85;
    --tm-border:              #DAD7D0;
    --tm-border-hover:        #B8B5AD;
    --tm-border-strong:       #1A1A1A;
    --tm-success:             #4F7A3A;
    --tm-info:                #345878;
    --tm-warning:             #C8541C;
    --tm-destructive:         #A33B2A;
    --TM-font-serif: "Roboto Slab", Georgia, serif;
    --TM-font-sans:  "Inter", system-ui, sans-serif;
    --TM-font-mono:  "JetBrains Mono", ui-monospace, monospace;
    --TM-radius-sm: 4px;
    --TM-radius-md: 6px;
    --TM-radius-lg: 8px;
    --TM-radius-xl: 12px;
    --TM-shadow-sm: 0 1px 0 rgba(26,26,26,.04), 0 1px 2px rgba(26,26,26,.04);
    --TM-shadow-md: 0 2px 4px -1px rgba(26,26,26,.08), 0 1px 2px rgba(26,26,26,.04);
    --TM-shadow-lg: 0 8px 16px -4px rgba(26,26,26,.10), 0 2px 4px -1px rgba(26,26,26,.06);
  }
  html.dark {
    --TM-primary:             #E47A4A;
    --TM-primary-hover:       #C8541C;
    --TM-primary-light:       #2B2520;
    --TM-primary-light-alpha: rgba(228,122,74,0.18);
    --tm-bg:                  #161513;
    --tm-bg-card:             #1F1E1B;
    --tm-bg-hover:            #2A2926;
    --tm-bg-input:            #1F1E1B;
    --tm-bg-sidebar:          #1A1916;
    --tm-bg-secondary:        #2A2926;
    --tm-bg-accent:           #2B2520;
    --tm-bg-muted:            #232220;
    --tm-text:                #EFEDE8;
    --tm-text-muted:          #A8A6A0;
    --tm-text-subtle:         #6E6C66;
    --tm-border:              #34322E;
    --tm-border-hover:        #4A4842;
    --tm-border-strong:       #EFEDE8;
    --TM-shadow-sm: 0 1px 0 rgba(0,0,0,.40);
    --TM-shadow-md: 0 2px 4px -1px rgba(0,0,0,.50);
    --TM-shadow-lg: 0 8px 16px -4px rgba(0,0,0,.55);
  }

  body {
    font-family: var(--TM-font-sans);
    background: var(--tm-bg);
    color: var(--tm-text);
    font-size: 14px;
    line-height: 1.6;
    transition: background 0.3s ease, color 0.3s ease;
  }

  /* ── HEADER ── */
  header {
    background: var(--tm-bg-card);
    border-bottom: 1px solid var(--tm-border);
    padding: 0 24px;
    height: 56px;
    display: flex;
    align-items: center;
    gap: 16px;
    position: sticky;
    top: 0;
    z-index: 100;
    box-shadow: var(--TM-shadow-sm);
  }
  .hdr-brand {
    display: flex;
    align-items: center;
    gap: 10px;
    flex-shrink: 0;
  }
  .hdr-logo-mark {
    width: 28px; height: 28px;
    background: var(--TM-primary);
    border-radius: var(--TM-radius-md);
    display: flex; align-items: center; justify-content: center;
    flex-shrink: 0;
  }
  .hdr-logo-mark svg { width: 16px; height: 16px; stroke: #fff; fill: none; stroke-width: 2.5; stroke-linecap: round; stroke-linejoin: round; }
  .hdr-wordmark {
    font-family: var(--TM-font-serif);
    font-weight: 600;
    font-size: 14px;
    color: var(--tm-text);
    letter-spacing: -0.01em;
    white-space: nowrap;
  }
  .hdr-div { width: 1px; height: 20px; background: var(--tm-border); flex-shrink: 0; }
  .hdr-contract-badge {
    font-family: var(--TM-font-mono);
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: var(--TM-primary);
    background: var(--TM-primary-light);
    padding: 3px 9px;
    border-radius: var(--TM-radius-sm);
    border: 1px solid var(--TM-primary-light-alpha);
    white-space: nowrap;
  }
  .hdr-meta-text {
    font-family: var(--TM-font-sans);
    font-size: 12px;
    color: var(--tm-text-subtle);
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
  }
  .hdr-right {
    margin-left: auto;
    display: flex;
    align-items: center;
    gap: 8px;
    flex-shrink: 0;
  }
  .kbd-pill {
    font-family: var(--TM-font-mono);
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 0.08em;
    color: var(--tm-text-subtle);
    border: 1px solid var(--tm-border);
    background: var(--tm-bg);
    padding: 2px 8px;
    border-radius: var(--TM-radius-sm);
    white-space: nowrap;
  }
  .dark-toggle {
    width: 32px; height: 32px;
    border: 1px solid var(--tm-border);
    border-radius: var(--TM-radius-md);
    background: var(--tm-bg);
    cursor: pointer;
    display: flex; align-items: center; justify-content: center;
    transition: background 0.15s, border-color 0.15s;
    flex-shrink: 0;
  }
  .dark-toggle:hover { background: var(--tm-bg-hover); border-color: var(--tm-border-hover); }
  .dark-toggle svg { width: 15px; height: 15px; stroke: var(--tm-text-muted); fill: none; stroke-width: 1.75; stroke-linecap: round; stroke-linejoin: round; }

  /* ── botão de toggle da sidebar — fica no header, sempre visível ── */
  .sidebar-toggle {
    width: 32px; height: 32px;
    border: 1px solid var(--tm-border);
    border-radius: var(--TM-radius-md);
    background: var(--tm-bg);
    cursor: pointer;
    display: flex; align-items: center; justify-content: center;
    transition: background 0.15s, border-color 0.15s;
    flex-shrink: 0;
  }
  .sidebar-toggle:hover { background: var(--tm-bg-hover); border-color: var(--tm-border-hover); }
  .sidebar-toggle svg { width: 15px; height: 15px; stroke: var(--tm-text-muted); fill: none; stroke-width: 1.75; stroke-linecap: round; stroke-linejoin: round; transition: transform 0.22s ease; }
  .sidebar-toggle.collapsed svg { transform: rotate(180deg); }

  /* ── LAYOUT ── */
  .layout { display: flex; height: calc(100vh - 56px); overflow: hidden; }

  /* ── SIDEBAR ── */
  aside {
    width: 196px;
    background: var(--tm-bg-sidebar);
    border-right: 1px solid var(--tm-border);
    padding: 16px 0;
    flex-shrink: 0;
    overflow-y: auto;
    overflow-x: hidden;
    display: flex;
    flex-direction: column;
    transition: width 0.22s ease, padding 0.22s ease;
  }
  aside.collapsed { width: 0; padding: 0; border-right-color: transparent; }
  aside.collapsed > * { opacity: 0; pointer-events: none; transition: opacity 0.1s; }
  aside > * { transition: opacity 0.2s 0.05s; }
  .aside-label {
    font-family: var(--TM-font-mono);
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: var(--tm-text-subtle);
    padding: 0 14px 8px;
  }
  .un-btn {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 7px 14px;
    cursor: pointer;
    font-family: var(--TM-font-sans);
    font-size: 13px;
    font-weight: 400;
    border: none;
    background: none;
    width: 100%;
    text-align: left;
    color: var(--tm-text-muted);
    position: relative;
    transition: background 0.12s, color 0.12s;
  }
  .un-btn:hover { background: var(--tm-bg-hover); color: var(--tm-text); }
  .un-btn.active {
    background: var(--TM-primary-light);
    color: var(--TM-primary);
    font-weight: 600;
  }
  .un-btn.active::before {
    content: '';
    position: absolute;
    left: 0; top: 0; bottom: 0;
    width: 2px;
    background: var(--TM-primary);
  }
  .un-badge {
    margin-left: auto;
    font-family: var(--TM-font-mono);
    font-size: 10px;
    font-weight: 600;
    padding: 1px 6px;
    border-radius: var(--TM-radius-sm);
    background: var(--tm-bg-secondary);
    color: var(--tm-text-subtle);
  }
  .un-btn.active .un-badge {
    background: var(--TM-primary-light-alpha);
    color: var(--TM-primary);
  }
  .aside-sep { height: 1px; background: var(--tm-border); margin: 10px 14px; }
  .aside-section-label {
    font-family: var(--TM-font-mono);
    font-size: 10px;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: var(--tm-text-subtle);
    padding: 12px 14px 6px;
    font-weight: 500;
  }

  /* ── MAIN ── */
  main { flex: 1; overflow-y: auto; display: flex; flex-direction: column; }

  /* ── SEARCH ── */
  .search-wrap {
    position: sticky;
    top: 0;
    z-index: 50;
    background: var(--tm-bg);
    border-bottom: 1px solid var(--tm-border);
    padding: 12px 20px 10px;
    display: flex;
    flex-direction: column;
    gap: 8px;
  }
  .search-row { display: flex; gap: 8px; align-items: center; }
  .search-box { flex: 1; position: relative; }
  .search-icon {
    position: absolute; left: 11px; top: 50%; transform: translateY(-50%);
    pointer-events: none;
  }
  .search-icon svg { width: 15px; height: 15px; stroke: var(--tm-text-subtle); fill: none; stroke-width: 1.75; stroke-linecap: round; stroke-linejoin: round; }
  .search-box input {
    width: 100%;
    padding: 8px 34px 8px 34px;
    border: 1px solid var(--tm-border);
    border-radius: var(--TM-radius-md);
    font-family: var(--TM-font-sans);
    font-size: 13px;
    color: var(--tm-text);
    background: var(--tm-bg-input);
    outline: none;
    transition: border-color 0.15s, box-shadow 0.15s;
  }
  .search-box input::placeholder { color: var(--tm-text-subtle); }
  .search-box input:focus {
    border-color: var(--TM-primary);
    box-shadow: 0 0 0 3px var(--TM-primary-light-alpha);
  }
  .search-clear {
    position: absolute; right: 9px; top: 50%; transform: translateY(-50%);
    background: none; border: none; cursor: pointer;
    display: none; padding: 2px;
    border-radius: var(--TM-radius-sm);
    color: var(--tm-text-subtle);
  }
  .search-clear svg { width: 13px; height: 13px; stroke: currentColor; fill: none; stroke-width: 2; stroke-linecap: round; }
  .search-clear:hover { color: var(--tm-text); background: var(--tm-bg-hover); }
  .search-clear.visible { display: flex; align-items: center; }
  .cnt-badge {
    font-family: var(--TM-font-mono);
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.06em;
    color: var(--TM-primary);
    background: var(--TM-primary-light);
    border: 1px solid var(--TM-primary-light-alpha);
    padding: 4px 12px;
    border-radius: var(--TM-radius-sm);
    white-space: nowrap;
  }

  /* AUTOCOMPLETE */
  .suggestions {
    position: absolute;
    top: calc(100% + 4px); left: 0; right: 0;
    background: var(--tm-bg-card);
    border: 1px solid var(--tm-border);
    border-radius: var(--TM-radius-lg);
    box-shadow: var(--TM-shadow-lg);
    z-index: 200;
    max-height: 280px;
    overflow-y: auto;
    display: none;
  }
  .suggestions.open { display: block; }
  .sug-section-hdr {
    padding: 6px 14px 4px;
    font-family: var(--TM-font-mono);
    font-size: 10px;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: var(--tm-text-subtle);
    background: var(--tm-bg-muted);
    border-bottom: 1px solid var(--tm-border);
    font-weight: 500;
  }
  .sug-item {
    padding: 9px 14px;
    cursor: pointer;
    display: flex;
    align-items: center;
    gap: 10px;
    border-bottom: 1px solid var(--tm-border);
    transition: background 0.1s;
  }
  .sug-item:last-child { border-bottom: none; }
  .sug-item:hover, .sug-item.focused { background: var(--tm-bg-hover); }
  .sug-code {
    font-family: var(--TM-font-mono);
    font-weight: 600;
    font-size: 12px;
    color: var(--TM-primary);
    min-width: 36px;
    white-space: nowrap;
  }
  .sug-desc { font-size: 13px; color: var(--tm-text); flex: 1; line-height: 1.4; }
  .sug-un {
    font-family: var(--TM-font-mono);
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 0.08em;
    color: var(--tm-info);
    background: rgba(52,88,120,0.10);
    padding: 1px 6px;
    border-radius: var(--TM-radius-sm);
    white-space: nowrap;
  }
  .sug-pref {
    font-family: var(--TM-font-mono);
    font-size: 10px;
    font-weight: 600;
    color: var(--tm-success);
    background: rgba(79,122,58,0.10);
    padding: 1px 6px;
    border-radius: var(--TM-radius-sm);
  }
  mark {
    background: #FDE68A;
    color: #92400E;
    border-radius: 2px;
    padding: 0 1px;
    font-style: normal;
  }

  /* ── CONTENT ── */
  .content { padding: 16px 20px; flex: 1; }

  .secao {
    margin-bottom: 16px;
    border-radius: var(--TM-radius-lg);
    overflow: hidden;
    border: 1px solid var(--tm-border);
    box-shadow: var(--TM-shadow-sm);
    background: var(--tm-bg-card);
  }
  .secao-header {
    background: var(--tm-bg-secondary);
    border-bottom: 1px solid var(--tm-border);
    padding: 9px 16px;
    display: flex;
    align-items: center;
    gap: 10px;
  }
  .secao-num {
    font-family: var(--TM-font-mono);
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 0.12em;
    color: var(--TM-primary);
    background: var(--TM-primary-light);
    border: 1px solid var(--TM-primary-light-alpha);
    padding: 2px 8px;
    border-radius: var(--TM-radius-sm);
  }
  .secao-nome {
    font-family: var(--TM-font-sans);
    font-size: 12px;
    font-weight: 600;
    color: var(--tm-text);
    letter-spacing: 0.02em;
    text-transform: uppercase;
  }
  .secao-cnt {
    margin-left: auto;
    font-family: var(--TM-font-mono);
    font-size: 10px;
    color: var(--tm-text-subtle);
    font-weight: 500;
  }

  table { width: 100%; border-collapse: collapse; }
  th {
    font-family: var(--TM-font-mono);
    font-size: 10px;
    font-weight: 500;
    letter-spacing: 0.10em;
    text-transform: uppercase;
    color: var(--tm-text-subtle);
    padding: 8px 14px;
    text-align: left;
    background: var(--tm-bg-muted);
    border-bottom: 1px solid var(--tm-border);
  }
  td {
    padding: 9px 14px;
    font-size: 13px;
    border-bottom: 1px solid var(--tm-border);
    color: var(--tm-text);
    vertical-align: middle;
  }
  tr:last-child td { border-bottom: none; }
  tr:hover td { background: var(--tm-bg-hover); }

  .item-code {
    font-family: var(--TM-font-mono);
    font-weight: 600;
    font-size: 12px;
    color: var(--TM-primary);
    white-space: nowrap;
  }
  .item-desc { font-family: var(--TM-font-sans); font-size: 13px; color: var(--tm-text); }

  /* chips de unidade — paleta TM chart ramp */
  .un-chip {
    display: inline-block;
    font-family: var(--TM-font-mono);
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 0.08em;
    padding: 2px 8px;
    border-radius: var(--TM-radius-sm);
    white-space: nowrap;
  }
  .un-m2  { background: rgba(52,88,120,0.12); color: var(--tm-info); border: 1px solid rgba(52,88,120,0.18); }
  .un-m3  { background: rgba(163,59,42,0.10); color: var(--tm-destructive); border: 1px solid rgba(163,59,42,0.18); }
  .un-m   { background: rgba(79,122,58,0.10); color: var(--tm-success); border: 1px solid rgba(79,122,58,0.18); }
  .un-un  { background: var(--TM-primary-light); color: var(--TM-primary); border: 1px solid var(--TM-primary-light-alpha); }
  .un-km  { background: rgba(140,90,46,0.10); color: #8C5A2E; border: 1px solid rgba(140,90,46,0.18); }
  .un-cj  { background: rgba(43,43,43,0.07); color: var(--TM-accent-graphite); border: 1px solid rgba(43,43,43,0.15); }
  .un-mes { background: var(--tm-bg-secondary); color: var(--tm-text-muted); border: 1px solid var(--tm-border); }
  .un-h   { background: rgba(79,122,58,0.07); color: #4F7A3A; border: 1px solid rgba(79,122,58,0.14); }
  .un-def { background: var(--tm-bg-muted); color: var(--tm-text-subtle); border: 1px solid var(--tm-border); }

  /* ── NO RESULTS ── */
  .no-results {
    display: flex; flex-direction: column; align-items: center; gap: 10px;
    padding: 64px 20px; text-align: center;
  }
  .no-results-icon {
    width: 40px; height: 40px;
    stroke: var(--tm-text-subtle); fill: none; stroke-width: 1.5; stroke-linecap: round; stroke-linejoin: round;
  }
  .no-results-title { font-family: var(--TM-font-serif); font-size: 15px; font-weight: 500; color: var(--tm-text-muted); }
  .no-results-hint { font-size: 12px; color: var(--tm-text-subtle); }

  /* ── PREFIXOS ── */
  .pref-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(260px, 1fr));
    gap: 8px;
  }
  .pref-card {
    background: var(--tm-bg-card);
    border: 1px solid var(--tm-border);
    border-radius: var(--TM-radius-lg);
    padding: 10px 12px;
    display: flex;
    gap: 10px;
    align-items: center;
    transition: border-color 0.15s, box-shadow 0.15s;
  }
  .pref-card:hover { border-color: var(--tm-border-hover); box-shadow: var(--TM-shadow-md); }
  .pref-num {
    font-family: var(--TM-font-mono);
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.06em;
    color: var(--TM-primary);
    background: var(--TM-primary-light);
    border: 1px solid var(--TM-primary-light-alpha);
    padding: 4px 10px;
    border-radius: var(--TM-radius-sm);
    white-space: nowrap;
    min-width: 52px;
    text-align: center;
  }
  .pref-dep { font-weight: 600; font-size: 13px; color: var(--tm-text); }
  .pref-cidade { font-size: 11px; color: var(--tm-text-subtle); margin-top: 1px; }
  .pref-km {
    font-family: var(--TM-font-mono);
    font-size: 11px;
    font-weight: 400;
    padding: 3px 8px;
    border-radius: var(--TM-radius-sm);
    white-space: nowrap;
    flex-shrink: 0;
    align-self: center;
    display: flex;
    align-items: center;
    gap: 4px;
    border: 1px solid rgba(79,122,58,0.25);
    background: rgba(79,122,58,0.08);
  }
  .pref-km-label {
    color: var(--tm-text-subtle);
    font-weight: 600;
    letter-spacing: 0.06em;
    font-size: 10px;
    text-transform: uppercase;
  }
  .pref-km-val {
    color: var(--tm-success);
    font-weight: 500;
  }

  /* ── BUSCA ATIVA — oculta lista de seções e mostra só resultados ── */
  .searching .secao { display: none; }
  .searching .secao.tem-resultado { display: block; }

  /* ── SCROLL BAR ── */
  ::-webkit-scrollbar { width: 5px; height: 5px; }
  ::-webkit-scrollbar-track { background: transparent; }
  ::-webkit-scrollbar-thumb { background: var(--tm-border); border-radius: 10px; }
  ::-webkit-scrollbar-thumb:hover { background: var(--tm-border-hover); }

  :focus-visible { outline: 2px solid var(--TM-primary); outline-offset: 2px; }
  @media (prefers-reduced-motion: reduce) { *, *::before, *::after { transition-duration: 0.01ms !important; } }
  .hidden { display: none !important; }

  /* ── RESPONSIVO MOBILE ── */
  @media (max-width: 640px) {
    header { padding: 0 12px; gap: 8px; }
    .hdr-wordmark { display: none; }
    .hdr-meta-text { display: none; }
    .kbd-pill { display: none; }

    aside {
      position: fixed;
      top: 56px; left: 0; bottom: 0;
      z-index: 90;
      width: 220px;
      box-shadow: var(--TM-shadow-lg);
    }
    aside.collapsed { width: 0; box-shadow: none; }
    aside.collapsed .sidebar-toggle { right: -40px; }

    /* overlay escuro ao abrir sidebar no mobile */
    .sidebar-overlay {
      display: block;
      position: fixed;
      inset: 0;
      background: rgba(0,0,0,0.35);
      z-index: 89;
      opacity: 1;
      transition: opacity 0.22s;
    }
    .sidebar-overlay.hidden { opacity: 0; pointer-events: none; }

    .layout { height: calc(100vh - 56px); }
    main { width: 100%; }
    .search-wrap { padding: 10px 12px 8px; }
    .content { padding: 12px; }
    .pref-grid { grid-template-columns: 1fr; }
    table { font-size: 12px; }
    td, th { padding: 7px 10px; }
  }
  @media (min-width: 641px) {
    .sidebar-overlay { display: none !important; }
  }

  /* ── FOOTER ── */
  footer {
    text-align: center;
    padding: 12px 20px;
    font-family: var(--TM-font-mono);
    font-size: 10px;
    letter-spacing: 0.10em;
    text-transform: uppercase;
    color: var(--tm-text-subtle);
    border-top: 1px solid var(--tm-border);
    background: var(--tm-bg-sidebar);
    flex-shrink: 0;
  }
  footer span { color: var(--TM-primary); font-weight: 600; }
</style>
</head>
<body>
<header>
  <div class="hdr-brand">
    <div class="hdr-logo-mark">
      <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M3 9l9-7 9 7v11a2 2 0 01-2 2H5a2 2 0 01-2-2z"/><polyline points="9 22 9 12 15 12 15 22"/></svg>
    </div>
    <span class="hdr-wordmark">MAFFENG</span>
  </div>
  <div class="hdr-div"></div>
  <span class="hdr-contract-badge" id="hdr-codigo">—</span>
  <span class="hdr-meta-text" id="hdr-meta"></span>
  <div class="hdr-right">
    <span class="kbd-pill">Ctrl+K</span>
    <button class="sidebar-toggle" id="sidebar-toggle" onclick="toggleSidebar()" title="Menu lateral" aria-label="Abrir/fechar menu lateral">
      <svg viewBox="0 0 24 24" aria-hidden="true"><line x1="3" y1="6" x2="21" y2="6"/><line x1="3" y1="12" x2="21" y2="12"/><line x1="3" y1="18" x2="21" y2="18"/></svg>
    </button>
    <button class="dark-toggle" id="dark-toggle" title="Alternar tema" onclick="toggleDark()">
      <svg id="icon-moon" viewBox="0 0 24 24"><path d="M21 12.79A9 9 0 1111.21 3 7 7 0 0021 12.79z"/></svg>
      <svg id="icon-sun" viewBox="0 0 24 24" style="display:none"><circle cx="12" cy="12" r="5"/><line x1="12" y1="1" x2="12" y2="3"/><line x1="12" y1="21" x2="12" y2="23"/><line x1="4.22" y1="4.22" x2="5.64" y2="5.64"/><line x1="18.36" y1="18.36" x2="19.78" y2="19.78"/><line x1="1" y1="12" x2="3" y2="12"/><line x1="21" y1="12" x2="23" y2="12"/><line x1="4.22" y1="19.78" x2="5.64" y2="18.36"/><line x1="18.36" y1="5.64" x2="19.78" y2="4.22"/></svg>
    </button>
  </div>
</header>

<div class="sidebar-overlay hidden" id="sidebar-overlay" onclick="toggleSidebar()"></div>
<div class="layout">
  <aside id="aside">
    <div class="aside-label">Unidade</div>
    <button class="un-btn active" data-un="todos" onclick="filtrar('todos', true)">
      <span>Todos</span><span class="un-badge" id="cnt-todos">0</span>
    </button>
    <div class="aside-sep"></div>
    <div id="un-buttons"></div>
    <div class="aside-sep"></div>
    <div class="aside-section-label">Contrato</div>
    <button class="un-btn" data-un="prefixos" onclick="filtrar('prefixos', true)">
      <span>Agências</span><span class="un-badge" id="cnt-pref">0</span>
    </button>
  </aside>

  <main>
    <div class="search-wrap">
      <div class="search-row">
        <div class="search-box">
          <span class="search-icon">
            <svg viewBox="0 0 24 24" aria-hidden="true"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>
          </span>
          <input type="text" id="busca"
            placeholder="Buscar por código (2.12), palavra-chave (piso), sigla (ART)..."
            autocomplete="off" spellcheck="false">
          <button class="search-clear" id="btn-clear" onclick="limparBusca()" title="Limpar (Esc)">
            <svg viewBox="0 0 24 24" aria-hidden="true"><line x1="18" y1="6" x2="6" y2="18"/><line x1="6" y1="6" x2="18" y2="18"/></svg>
          </button>
          <div class="suggestions" id="suggestions"></div>
        </div>
        <div class="cnt-badge" id="cnt-exibidos">0 itens</div>
      </div>
    </div>

    <div class="content">
      <div id="itens-view"></div>
      <div id="prefixos-view" class="hidden"></div>
    </div>
    <footer>Feito por <span>TM Sempre Tecnologia</span></footer>
  </main>
</div>

<script>
const DADOS = {{DADOS_JSON}};

// ── sidebar toggle ────────────────────────────────────────────────────
function toggleSidebar() {
  const aside   = document.getElementById('aside');
  const btn     = document.getElementById('sidebar-toggle');
  const overlay = document.getElementById('sidebar-overlay');
  const collapsed = aside.classList.toggle('collapsed');
  btn.classList.toggle('collapsed', collapsed);
  // overlay só no mobile
  if (window.innerWidth <= 640) {
    overlay.classList.toggle('hidden', collapsed);
  }
}

// ── dark mode ─────────────────────────────────────────────────────────
function toggleDark() {
  const dark = document.documentElement.classList.toggle('dark');
  localStorage.setItem('tm-theme', dark ? 'dark' : 'light');
  document.getElementById('icon-moon').style.display = dark ? 'none' : '';
  document.getElementById('icon-sun').style.display  = dark ? ''     : 'none';
}
(function applyTheme() {
  const saved = localStorage.getItem('tm-theme');
  const prefersDark = window.matchMedia('(prefers-color-scheme: dark)').matches;
  if (saved === 'dark' || (!saved && prefersDark)) {
    document.documentElement.classList.add('dark');
    document.getElementById('icon-moon').style.display = 'none';
    document.getElementById('icon-sun').style.display  = '';
  }
})();

// ── estado ──────────────────────────────────────────────────────────
let filtroAtual = 'todos';
let buscaAtual  = '';
let sugIdx      = -1;

// ── índice flat para autocomplete ───────────────────────────────────
const IDX = [];
DADOS.secoes.forEach(sec => {
  sec.itens.forEach(it => {
    IDX.push({ tipo: 'item', item: it.item, descricao: it.descricao, unidade: it.unidade, secao: sec.nome });
  });
});
DADOS.prefixos.forEach(p => {
  IDX.push({ tipo: 'prefixo', prefixo: p.prefixo, dependencia: p.dependencia, endereco: p.endereco || '', cidade: p.cidade || '' });
});

// ── normalização (remove acentos, lowercase) ────────────────────────
function norm(s) {
  return String(s).toLowerCase()
    .normalize('NFD').replace(/[̀-ͯ]/g, '')
    .replace(/[^a-z0-9\.\s]/g, ' ').replace(/\s+/g, ' ').trim();
}

// ── score de relevância (fuzzy simples por tokens) ───────────────────
function score(haystack, tokens) {
  const h = norm(haystack);
  let s = 0;
  for (const t of tokens) {
    if (h === t) { s += 100; continue; }
    if (h.startsWith(t)) { s += 60; continue; }
    if (h.includes(t)) { s += 30; continue; }
    // fuzzy: todos os chars do token aparecem em ordem?
    let pos = 0;
    let found = true;
    for (const c of t) { pos = h.indexOf(c, pos); if (pos < 0) { found = false; break; } pos++; }
    if (found && t.length >= 3) s += 10;
    else if (!found) s -= 20;
  }
  return s;
}

function scorarItem(it, tokens) {
  return Math.max(
    score(it.item, tokens) * 2,   // código tem peso dobrado
    score(it.descricao, tokens)
  );
}

// ── highlight de texto ────────────────────────────────────────────────
function hl(text, tokens) {
  if (!tokens || tokens.length === 0) return escHtml(text);
  let result = escHtml(text);
  tokens.forEach(t => {
    if (t.length < 2) return;
    const re = new RegExp('(' + t.replace(/[.*+?^${}()|[\]\\]/g, '\\$&') + ')', 'gi');
    result = result.replace(re, '<mark>$1</mark>');
  });
  return result;
}

function escHtml(s) {
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}

// ── classe CSS da unidade ─────────────────────────────────────────────
function unClass(u) {
  const map = { 'm²':'un-m2','m³':'un-m3','m':'un-m','un':'un-un','km':'un-km',
                'cj':'un-cj','m²/mês':'un-mes','m/mês':'un-mes','mês':'un-mes','H':'un-h' };
  return map[u] || 'un-def';
}

// ── INIT ──────────────────────────────────────────────────────────────
function init() {
  document.getElementById('hdr-codigo').textContent = DADOS.codigo + ' — ' + DADOS.nome;
  const desc = DADOS.desconto ? (DADOS.desconto * 100).toFixed(2) + '%' : '—';
  const bdi  = DADOS.bdi     ? (DADOS.bdi     * 100).toFixed(2) + '%' : '—';
  const metaParts = ['Contrato: ' + (DADOS.contrato || '—'), 'Lote: ' + (DADOS.lote || '—')];
  if (desc !== '—') metaParts.push('Desc: ' + desc);
  if (bdi  !== '—') metaParts.push('BDI: ' + bdi);
  document.getElementById('hdr-meta').textContent = metaParts.join('  ·  ');

  // Sidebar de unidades
  const unBtn = document.getElementById('un-buttons');
  const unOrdem = ['m²','m³','m','un','km','cj','m²/mês','m/mês','mês','H'];
  const unExist = Object.keys(DADOS.itens_por_unidade);
  const unOrd = [...unOrdem.filter(u => unExist.includes(u)), ...unExist.filter(u => !unOrdem.includes(u))];
  let tot = 0;
  unOrd.forEach(un => {
    const cnt = DADOS.itens_por_unidade[un].length;
    tot += cnt;
    const btn = document.createElement('button');
    btn.className = 'un-btn';
    btn.dataset.un = un;
    btn.onclick = () => filtrar(un, true);
    btn.innerHTML = `<span>${un}</span><span class="un-badge">${cnt}</span>`;
    unBtn.appendChild(btn);
  });
  document.getElementById('cnt-todos').textContent = tot;
  document.getElementById('cnt-pref').textContent = DADOS.prefixos.length;

  renderPrefixos();
  renderItens([]);

  // Input events
  const inp = document.getElementById('busca');
  inp.addEventListener('input', onInput);
  inp.addEventListener('keydown', onKeyDown);
  inp.addEventListener('blur', () => setTimeout(() => fecharSug(), 180));
  inp.addEventListener('focus', () => { if (inp.value) onInput(); });

  // Atalho global Ctrl+K
  document.addEventListener('keydown', e => {
    if ((e.ctrlKey || e.metaKey) && e.key === 'k') { e.preventDefault(); inp.focus(); inp.select(); }
    if (e.key === 'Escape') { limparBusca(); inp.blur(); }
  });
}

// ── FILTRO SIDEBAR ────────────────────────────────────────────────────
function filtrar(un, limparQ) {
  filtroAtual = un;
  if (limparQ) { limparBusca(); }
  document.querySelectorAll('.un-btn').forEach(b => b.classList.remove('active'));
  document.querySelector(`[data-un="${CSS.escape(un)}"]`)?.classList.add('active');

  const itV = document.getElementById('itens-view');
  const prV = document.getElementById('prefixos-view');
  if (un === 'prefixos') {
    itV.classList.add('hidden'); prV.classList.remove('hidden');
    document.getElementById('cnt-exibidos').textContent = DADOS.prefixos.length + ' agências';
    return;
  }
  itV.classList.remove('hidden'); prV.classList.add('hidden');
  renderItens(tokenizar(buscaAtual));
}

// ── INPUT HANDLER ─────────────────────────────────────────────────────
let _sugTimer = null;
function onInput() {
  const v = document.getElementById('busca').value;
  buscaAtual = v;
  document.getElementById('btn-clear').classList.toggle('visible', v.length > 0);

  const tokens = tokenizar(v);

  if (v.trim().length >= 2) {
    // Esconde a tabela e agenda o dropdown
    document.getElementById('itens-view').style.visibility = 'hidden';
    document.getElementById('suggestions').classList.remove('open');
    clearTimeout(_sugTimer);
    _sugTimer = setTimeout(() => {
      if (document.activeElement === document.getElementById('busca')) {
        renderSugestoes(tokens, v);
      }
    }, 300);
  } else {
    // Campo vazio ou 1 char: mostra tudo normal
    fecharSug();
    renderItens(tokens);
  }
}

function tokenizar(v) {
  return norm(v).split(' ').filter(t => t.length >= 1);
}

// ── SUGESTÕES ────────────────────────────────────────────────────────
function renderSugestoes(tokens, raw) {
  const sug = document.getElementById('suggestions');
  sugIdx = -1;
  document.getElementById('itens-view').style.visibility = 'hidden';

  // Pontua e ordena
  const resultados = IDX
    .map(entry => {
      let s = 0;
      if (entry.tipo === 'item') {
        s = scorarItem(entry, tokens);
      } else {
        s = Math.max(score(entry.prefixo, tokens) * 2, score(entry.dependencia, tokens), score(entry.cidade, tokens), score(entry.endereco, tokens));
      }
      return { entry, s };
    })
    .filter(x => x.s > 0)
    .sort((a, b) => b.s - a.s)
    .slice(0, 12);

  if (resultados.length === 0) { fecharSug(); return; }

  let html = '';
  let lastTipo = null;
  resultados.forEach(({ entry }) => {
    if (entry.tipo !== lastTipo) {
      html += `<div class="sug-section-hdr">${entry.tipo === 'item' ? 'Itens' : 'Agências'}</div>`;
      lastTipo = entry.tipo;
    }
    if (entry.tipo === 'item') {
      html += `<div class="sug-item" onclick="selecionarItem('${escHtml(entry.item)}')">
        <span class="sug-code">${hl(entry.item, tokens)}</span>
        <span class="sug-desc">${hl(entry.descricao, tokens)}</span>
        <span class="sug-un">${entry.unidade}</span>
      </div>`;
    } else {
      html += `<div class="sug-item" onclick="selecionarPrefixo('${escHtml(entry.prefixo)}')">
        <span class="sug-code">${hl(entry.prefixo, tokens)}</span>
        <span class="sug-desc">${hl(entry.dependencia, tokens)}</span>
        <span class="sug-pref">${hl(entry.cidade, tokens)}</span>
      </div>`;
    }
  });

  sug.innerHTML = html;
  sug.classList.add('open');
}

function fecharSug() {
  document.getElementById('suggestions').classList.remove('open');
  document.getElementById('itens-view').style.visibility = '';
  sugIdx = -1;
}

function selecionarItem(codigo) {
  document.getElementById('busca').value = codigo;
  buscaAtual = codigo;
  document.getElementById('btn-clear').classList.add('visible');
  fecharSug();
  document.getElementById('itens-view').style.visibility = '';
  renderItens(tokenizar(codigo));
}

function selecionarPrefixo(pref) {
  filtrar('prefixos', false);
  document.getElementById('busca').value = pref;
  buscaAtual = pref;
  fecharSug();
  renderPrefixos(pref);
}

// ── TECLADO nas sugestões ──────────────────────────────────────────────
function onKeyDown(e) {
  const sug = document.getElementById('suggestions');
  const items = sug.querySelectorAll('.sug-item');
  if (!sug.classList.contains('open')) return;

  if (e.key === 'ArrowDown') {
    e.preventDefault();
    sugIdx = Math.min(sugIdx + 1, items.length - 1);
    items.forEach((el, i) => el.classList.toggle('focused', i === sugIdx));
    items[sugIdx]?.scrollIntoView({ block: 'nearest' });
  } else if (e.key === 'ArrowUp') {
    e.preventDefault();
    sugIdx = Math.max(sugIdx - 1, 0);
    items.forEach((el, i) => el.classList.toggle('focused', i === sugIdx));
    items[sugIdx]?.scrollIntoView({ block: 'nearest' });
  } else if (e.key === 'Enter' && sugIdx >= 0) {
    e.preventDefault();
    items[sugIdx]?.click();
  } else if (e.key === 'Escape') {
    fecharSug();
  }
}

// ── LIMPAR ────────────────────────────────────────────────────────────
function limparBusca() {
  document.getElementById('busca').value = '';
  buscaAtual = '';
  document.getElementById('btn-clear').classList.remove('visible');
  fecharSug();
  renderItens([]);
  if (filtroAtual === 'prefixos') renderPrefixos('');
}

// ── RENDER ITENS ──────────────────────────────────────────────────────
function itemPassaFiltro(it, tokens) {
  if (filtroAtual !== 'todos' && filtroAtual !== 'prefixos' && it.unidade !== filtroAtual) return false;
  if (tokens && tokens.length > 0) {
    return scorarItem(it, tokens) > 0;
  }
  return true;
}

function renderItens(tokens) {
  if (filtroAtual === 'prefixos') return;
  const view = document.getElementById('itens-view');
  const buscando = tokens && tokens.length > 0;
  let total = 0;
  let html = '';

  // Ordena seções por score quando há busca
  const secoesComScore = DADOS.secoes.map(sec => {
    const itensFilt = sec.itens.filter(it => itemPassaFiltro(it, tokens));
    const maxScore = buscando
      ? Math.max(0, ...itensFilt.map(it => scorarItem(it, tokens)))
      : 0;
    return { sec, itensFilt, maxScore };
  });

  // Quando buscando: ordena por score e mostra só quem tem resultado
  const secoesVisiveis = buscando
    ? [...secoesComScore].filter(x => x.itensFilt.length > 0).sort((a, b) => b.maxScore - a.maxScore)
    : secoesComScore.filter(x => x.itensFilt.length > 0);

  secoesVisiveis.forEach(({ sec, itensFilt }) => {
    total += itensFilt.length;

    const itensOrdenados = buscando
      ? [...itensFilt].sort((a, b) => scorarItem(b, tokens) - scorarItem(a, tokens))
      : itensFilt;

    html += `<div class="secao">
      <div class="secao-header">
        <span class="secao-num">${escHtml(sec.codigo)}</span>
        <span class="secao-nome">${escHtml(sec.nome)}</span>
        <span class="secao-cnt">${itensFilt.length} iten${itensFilt.length > 1 ? 's' : ''}</span>
      </div>
      <table>
        <thead><tr>
          <th style="width:78px">Item</th>
          <th>Descrição</th>
          <th style="width:88px;text-align:center">UN</th>
        </tr></thead><tbody>`;

    itensOrdenados.forEach(it => {
      const hlItem = hl(it.item, tokens);
      const hlDesc = hl(it.descricao, tokens);
      html += `<tr>
        <td class="item-code">${hlItem}</td>
        <td>${hlDesc}</td>
        <td style="text-align:center">
          <span class="un-chip ${unClass(it.unidade)}">${it.unidade}</span>
        </td>
      </tr>`;
    });

    html += `</tbody></table></div>`;
  });

  if (total === 0) {
    const q = buscaAtual;
    html = `<div class="no-results">
      <svg class="no-results-icon" viewBox="0 0 24 24" aria-hidden="true"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/><line x1="11" y1="8" x2="11" y2="14"/><line x1="8" y1="11" x2="14" y2="11"/></svg>
      <div class="no-results-title">Nenhum item para <strong>"${escHtml(q)}"</strong></div>
      <div class="no-results-hint">Busque pelo código (ex: 2.12) ou parte da descrição (ex: piso tátil)</div>
    </div>`;
  }

  view.innerHTML = html;
  document.getElementById('cnt-exibidos').textContent = total + ' iten' + (total !== 1 ? 's' : '');
}

// ── RENDER PREFIXOS ───────────────────────────────────────────────────
function renderPrefixos(filtro) {
  const view = document.getElementById('prefixos-view');
  if (DADOS.prefixos.length === 0) {
    view.innerHTML = '<div class="no-results"><svg class="no-results-icon" viewBox="0 0 24 24" aria-hidden="true"><rect x="3" y="3" width="18" height="18" rx="2" ry="2"/><line x1="3" y1="9" x2="21" y2="9"/><line x1="9" y1="21" x2="9" y2="9"/></svg><div class="no-results-title">Sem agências cadastradas.</div></div>';
    return;
  }
  const q = norm(filtro || '');
  const tokens = q.split(' ').filter(t => t.length >= 1);
  const lista = q
    ? DADOS.prefixos.filter(p =>
        score(p.prefixo, tokens) > 0 || score(p.dependencia, tokens) > 0 || score(p.cidade || '', tokens) > 0 || score(p.endereco || '', tokens) > 0
      )
    : DADOS.prefixos;

  let html = `<div class="pref-grid">`;
  lista.forEach(p => {
    const partes = [];
    if (p.endereco) partes.push(p.endereco);
    if (p.cidade && p.cidade !== p.dependencia) partes.push(p.cidade);
    const linha2 = partes.join(' — ');
    html += `<div class="pref-card">
      <div class="pref-num">${hl(p.prefixo, tokens)}</div>
      <div style="flex:1;min-width:0">
        <div class="pref-dep">${hl(p.dependencia, tokens)}</div>
        ${linha2 ? `<div class="pref-cidade">${hl(linha2, tokens)}</div>` : ''}
      </div>
      ${p.km ? `<div class="pref-km"><span class="pref-km-label">km</span><span class="pref-km-val">${hl(p.km.replace(' km',''), tokens)}</span></div>` : ''}
    </div>`;
  });
  html += `</div>`;
  view.innerHTML = html;
  document.getElementById('cnt-exibidos').textContent = lista.length + ' agência' + (lista.length !== 1 ? 's' : '');
}

init();
</script>
</body>
</html>"""


def gerar_html(dados, titulo):
    dados_json = json.dumps(dados, ensure_ascii=False, indent=2)
    html = HTML_TEMPLATE.replace("{{TITULO}}", titulo)
    html = html.replace("{{DADOS_JSON}}", dados_json)
    return html


def main():
    print("=== Gerador de Memoriais de Itens — TM Sempre Tecnologia ===\n")
    resultados = []

    # Mapeia cada arquivo xlsx da raiz ao contrato correspondente
    # Chave: substring que identifica o contrato no nome do arquivo
    ARQUIVO_MAP = {
        "VARGINHA":      ("2057", "VARGINHA"),
        "SÃO PAULO":     ("0908", "SAO_PAULO"),
        "SAO PAULO":     ("0908", "SAO_PAULO"),
        "CUIABA":        ("1507", "CUIABA"),
        "SALINAS":       ("2626", "SALINAS"),
        "VALADARES":     ("2627", "VALADARES"),
        "TANGARA":       ("3575", "TANGARA"),
        "SP 1565":       ("1565", "SJRP"),
    }

    processados = set()
    for entry in sorted(os.scandir(BASE), key=lambda e: e.name):
        if not entry.name.lower().endswith(".xlsx"):
            continue
        # Ignora arquivos de saída ou auxiliares
        nome_upper = entry.name.upper()
        if "MEMORIAL DE ITENS" in nome_upper:
            continue

        codigo, slug = None, None
        for chave, (cod, sl) in ARQUIVO_MAP.items():
            if chave.upper() in nome_upper:
                codigo, slug = cod, sl
                break
        if not codigo or codigo in processados:
            continue

        processados.add(codigo)
        dados = processar_contrato(entry.path, codigo, slug)
        if not dados:
            continue

        # Salva JSON
        json_nome = f"{codigo}_{slug}_itens.json"
        json_path = os.path.join(BASE, json_nome)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
        print(f"    JSON salvo: {json_nome}")

        # Salva HTML
        titulo = f"{codigo} — {slug.replace('_',' ')} | Memorial de Itens"
        html_content = gerar_html(dados, titulo)
        html_nome = f"{codigo}_{slug}_memorial_itens.html"
        html_path = os.path.join(BASE, html_nome)
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        print(f"    HTML salvo: {html_nome}")

        resultados.append((codigo, slug, json_path, html_path))

    print(f"\n=== Concluído: {len(resultados)} contratos processados ===")
    for cod, slug, jp, hp in resultados:
        total = sum(len(s["itens"]) for s in json.load(open(jp, encoding="utf-8"))["secoes"])
        print(f"  {cod} {slug}: {total} itens")

if __name__ == "__main__":
    main()
